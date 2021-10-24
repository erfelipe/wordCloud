from openpyxl import load_workbook
from wordcloud import WordCloud 
import os
import re
import os.path

def repeteString(texto, n):
    """ Repete uma string n vezes e retorna sua concatenacao em nova string

    Args:
        texto (str): texto a ser repetido
        n (int): numero de vezes a ser repetida

    Returns:
        str: nova string com o numero de repeticao
    """    
    return ' '.join([texto for i in range(n)])

def gravarStringEmArquivo(texto, nomeArq):
    """ Grava variavel string em arquivo texto

    Args:
        texto (str): string a ser gravada em arquivo 
        nomeArq (str): nome do arquivo com path
    """    
    with open(nomeArq, 'w', encoding="utf-8") as text_file:
        text_file.write(texto)

def retirarExcessoDeEspacos(text):
    """ Retira dois ou mais espaços entre as palavras
    
    Arguments:
        text {str} -- Texto a ser tratado
    
    Returns:
        str -- Texto sem excesso de espaços entre as palavras 
    """    
    resp = re.sub("[ ]{2,}", " ", text)
    return resp

def arqsFromFolder(tpe):
    arqs = []
    for root, dirs, files in os.walk('Data'):
    # select file name
        for file in files:
            # check the extension of files
            if tpe == 'PLAN':
                if file.endswith('.xls') or file.endswith('.xlsx'):
                    # print whole path of files
                    arqs.append(os.path.join(root, file))
            elif tpe == 'TEXT':
                if file.endswith('.txt'):
                    # print whole path of files
                    arqs.append(os.path.join(root, file))
    return arqs

def madeTextFromArqs(arqs):
    for arq in arqs:
        wb = load_workbook(arq)
        ws = wb.active
        # print(wb.sheetnames)
        allWords = ""
        for row in ws.iter_rows(min_row=1, max_col=2, max_row=200, values_only=True):
            if (row[0] is not None):
                allWords += ' ' + repeteString(row[0], row[1]) 
        arqTxt = arq
        pre, ext = os.path.splitext(arq)
        arqTxt = pre + '.txt' 
        gravarStringEmArquivo(retirarExcessoDeEspacos(allWords).strip(), arqTxt)

def gerarNuvemDePalavrasPorArquivo(arqsTxt):
    """ Gera imagem com nuvem de palavras baseado em arquivo texto 
        Deve ser chamada depois do metodo separaStopWords() que gera o arquivo texto

    Args:
        nomeArq (str): Nome completo do arquivo (com path)
    """
    for nomeArq in arqsTxt:
        nomePng = nomeArq
        pre, ext = os.path.splitext(nomeArq)
        nomePng = pre + '.png' 

        # get data directory (using getcwd() is needed to support running example in generated IPython notebook)
        d = os.path.dirname(__file__) if "__file__" in locals() else os.getcwd()

        # Read the whole text.
        text = open(os.path.join(d, nomeArq)).read()

        # Generate a word cloud image
        wordcloud = WordCloud(width=1200, height=600, background_color="white", repeat=False, collocations=False, min_word_length=2)
        wordcloud.generate(text)
        wordcloud.to_file(nomePng)

if __name__ == "__main__":
    arqs = arqsFromFolder('PLAN')
    madeTextFromArqs(arqs)
    arqsTxt = arqsFromFolder('TEXT')
    gerarNuvemDePalavrasPorArquivo(arqsTxt)  